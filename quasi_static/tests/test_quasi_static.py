"""Tests for Updated Weight Method and quasi-static FSI smoke run."""

from __future__ import annotations

import sys
from pathlib import Path

import numpy as np
import pytest

QS = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(QS))

from membrane.geometry import build_rectangular_membrane
from utils.io import load_config
from uwm import compute_edge_weights, updated_weight_form_find
from coupling import QuasiStaticFSI


def test_uwm_weights_positive():
    mesh = build_rectangular_membrane(nx=4, ny=3, fixed_edges=["left", "right", "bottom", "top"])
    N_pre = 50.0  # N/m
    W = compute_edge_weights(mesh.nodes, mesh.elements, N_pre)
    assert len(W) > 0
    assert all(w > 0 for w in W.values())


def test_uwm_form_find_keeps_supports():
    mesh = build_rectangular_membrane(
        length=2.0, width=1.5, nx=6, ny=4,
        fixed_edges=["left", "right", "bottom", "top"],
    )
    x0 = mesh.nodes.copy()
    # perturb free nodes
    free = ~mesh.fixed
    mesh.nodes[free, 2] += 0.02 * np.sin(
        np.pi * (mesh.nodes[free, 0] - mesh.nodes[:, 0].min()) / mesh.length
    ) * np.sin(
        np.pi * (mesh.nodes[free, 1] - mesh.nodes[:, 1].min()) / mesh.width
    )
    res = updated_weight_form_find(
        mesh.nodes, mesh.elements, mesh.fixed, N_pre=50.0, max_weight_updates=15
    )
    assert np.isfinite(res.nodes).all()
    assert np.allclose(res.nodes[mesh.fixed], x0[mesh.fixed], atol=1e-12)
    # with isotropic prestress and planar supports, form stays near the plane
    assert np.max(np.abs(res.nodes[:, 2] - x0[:, 2])) < 0.05


def test_uwm_responds_to_pressure_load():
    mesh = build_rectangular_membrane(
        nx=6, ny=4, fixed_edges=["left", "right", "bottom", "top"]
    )
    f = np.zeros_like(mesh.nodes)
    f[~mesh.fixed, 2] = -2.0  # downward nodal loads
    res = updated_weight_form_find(
        mesh.nodes,
        mesh.elements,
        mesh.fixed,
        N_pre=40.0,
        f_ext=f,
        max_weight_updates=20,
    )
    assert res.nodes[~mesh.fixed, 2].mean() < -1e-4


def test_excel_deformations_export(tmp_path):
    from openpyxl import load_workbook
    from excel_export import DeformationRecorder

    mesh = build_rectangular_membrane(
        nx=4, ny=3, fixed_edges=["left", "right", "bottom", "top"]
    )
    ref = mesh.nodes.copy()
    rec = DeformationRecorder(reference_nodes=ref, fixed=mesh.fixed)
    for k in range(3):
        nodes = ref.copy()
        nodes[~mesh.fixed, 2] -= 0.01 * (k + 1)
        rec.record(time=0.1 * (k + 1), nodes=nodes, iteration=k + 1)
    path = rec.write_xlsx(tmp_path / "defs.xlsx", per_step_sheets=True)
    assert path.exists()
    wb = load_workbook(path)
    assert "summary" in wb.sheetnames
    assert "reference" in wb.sheetnames
    assert "deformations" in wb.sheetnames
    assert "step_0001" in wb.sheetnames
    # 3 steps × n_nodes data rows (+ header)
    ws = wb["deformations"]
    assert ws.max_row == 1 + 3 * mesh.n_nodes
    # spot-check: last free-node row of last step has uz ≈ -0.03
    free_ids = np.where(~mesh.fixed)[0]
    assert free_ids.size > 0
    # deformations rows are ordered by step then node_id
    # row index (1-based): 1 header + (step-1)*n_nodes + (nid+1)
    nid = int(free_ids[0])
    row_idx = 1 + (3 - 1) * mesh.n_nodes + (nid + 1)
    vals = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
    assert vals[0] == 3 and vals[3] == nid
    assert abs(vals[9] - (-0.03)) < 1e-12


def test_quasi_static_smoke():
    cfg = load_config(QS / "config" / "quasi_static.yaml")
    cfg["membrane"]["nx"] = 6
    cfg["membrane"]["ny"] = 4
    cfg["fluid"]["nx"] = 12
    cfg["fluid"]["ny"] = 6
    cfg["fluid"]["nz"] = 6
    cfg["fluid"]["nu"] = 5.0e-3
    cfg["time"]["dt"] = 0.01
    cfg["quasi_static"]["max_iters"] = 2
    cfg["quasi_static"]["fluid_substeps"] = 3
    cfg["quasi_static"]["load_scale"] = 0.1
    cfg["les"]["enabled"] = True
    sim = QuasiStaticFSI(cfg)
    hist = sim.run()
    assert len(hist.iteration) >= 1
    assert np.isfinite(sim.nodes).all()
    assert np.isfinite(sim.fluid.state.u).all()
    assert hist.shape_residual[-1] >= 0.0


def test_quasi_static_timed_gif_frame(tmp_path):
    from utils.viz import render_flutter_frame, save_gif

    cfg = load_config(QS / "config" / "quasi_static.yaml")
    cfg["membrane"]["nx"] = 6
    cfg["membrane"]["ny"] = 4
    cfg["fluid"]["nx"] = 12
    cfg["fluid"]["ny"] = 6
    cfg["fluid"]["nz"] = 6
    cfg["fluid"]["nu"] = 5.0e-3
    cfg["time"]["dt"] = 0.01
    cfg["time"]["t_end"] = 0.06
    cfg["quasi_static"]["fluid_substeps"] = 2
    cfg["quasi_static"]["max_iters"] = 20
    cfg["quasi_static"]["shape_tol"] = 0.0
    cfg["quasi_static"]["load_scale"] = 0.2
    cfg["les"]["enabled"] = True
    sim = QuasiStaticFSI(cfg)
    nodes0 = sim.x_bc.copy()
    frames = []
    j = sim.grid.ny // 2

    def on_frame(simulation, info, k):
        st = simulation.fluid.state
        speed = np.sqrt(
            st.u[:, j, :] ** 2 + st.v[:, j, :] ** 2 + st.w[:, j, :] ** 2
        )
        frames.append(
            render_flutter_frame(
                simulation.nodes,
                simulation.mesh.elements,
                nodes0,
                speed,
                simulation.grid.x,
                simulation.grid.z,
                info["time"],
                simulation.mesh.nx,
                simulation.mesh.ny,
                8.0,
                0.1,
                (0.7, 1.3),
                title="Quasi-static UWM membrane",
            )
        )

    sim.run_timed(t_end=0.06, callback=on_frame)
    assert len(frames) >= 2
    assert frames[-1] is not None
    gif = save_gif(frames, tmp_path / "qs.gif", fps=5)
    assert gif.exists() and gif.stat().st_size > 0
    assert sim.time >= 0.06 - 1e-12
